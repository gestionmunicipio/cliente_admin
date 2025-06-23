from django.shortcuts import render, redirect, get_object_or_404

from django.contrib.auth.decorators import login_required
from django.contrib.auth import login, logout, authenticate
from .models import Cliente, Direccion
from .forms import ClienteForm, DireccionFormSet, DireccionForm
from django.forms import inlineformset_factory

from django.contrib.auth.decorators import permission_required
from django.utils.translation import gettext_lazy as _
from django.db.models import Prefetch

### Exportar a Excel ####
import openpyxl
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
from clientes.models import Cliente

### Exportar a Pdf ####
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter


# Listar clientes
@login_required
@permission_required('clientes.view_cliente', raise_exception=True)
def lista_clientes(request):
    # DEBUG: imprime los permisos del usuario
    # print(request.user.get_all_permissions())
    
    clientes = Cliente.objects.all()
    return render(request, 'clientes/lista.html', {'clientes': clientes})

# Agregar cliente
@login_required
@permission_required('clientes.view_cliente', raise_exception=True)
#def agregar_cliente(request):
#    if request.method == 'POST':
#        form = ClienteForm(request.POST)
#        if form.is_valid():
#            form.save()
#            return redirect('lista_clientes')
#    else:
#        form = ClienteForm()
#    return render(request, 'clientes/formulario.html', {'form': form})

def agregar_cliente(request):
    if request.method == 'POST':
        cliente_form = ClienteForm(request.POST)
        if cliente_form.is_valid():
            cliente = cliente_form.save()
            direccion_formset = DireccionFormSet(request.POST, instance=cliente)
            if direccion_formset.is_valid():
                direccion_formset.save()
                return redirect('lista_clientes')
        else:
            direccion_formset = DireccionFormSet(request.POST)
    else:
        cliente_form = ClienteForm()
        direccion_formset = DireccionFormSet()

    return render(request, 'clientes/formulario.html', {
        'form': cliente_form,
        'direccion_formset': direccion_formset,
    })


# Editar cliente
@login_required
@permission_required('clientes.view_cliente', raise_exception=True)
def editar_cliente(request, pk):
    cliente = get_object_or_404(Cliente, pk=pk)
    if request.method == 'POST':
        form = ClienteForm(request.POST, instance=cliente)
        formset = DireccionFormSet(request.POST, instance=cliente)
        if form.is_valid():
            form.save()
            formset.save()
            return redirect('lista_clientes')
    else:
        form = ClienteForm(instance=cliente)
        formset = DireccionFormSet(instance=cliente)
    
    # Agregar las direcciones del cliente al contexto
    direcciones = cliente.direcciones.all()  # Solo si usaste related_name='direcciones'

    return render(request, 'clientes/formulario.html', {
        'form': form,
        'cliente': cliente,
        'direcciones': direcciones,
     #   'formset': formset,     # Puedes omitir si no lo usas directamente en el template
        'direccion_formset': formset,   # Necesario si lo usas como direccion_formset
    })
 
 
# Eliminar cliente
@login_required
def eliminar_cliente(request, pk):
    cliente = get_object_or_404(Cliente, pk=pk)
    if request.method == 'POST':
        cliente.delete()
        return redirect('lista_clientes')
    return render(request, 'clientes/confirmar_eliminar.html', {'cliente': cliente})

# Login
def login_view(request):
    error = None
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')
        user = authenticate(request, username=username, password=password)
        if user is not None:
            login(request, user)
            return redirect('lista_clientes')
        else:
            error = _("Usuario o contraseña incorrectos")
    return render(request, 'clientes/login.html', {'error': error})

# Logout
@login_required
def logout_view(request):
    logout(request)
    return redirect('login')

#===================================================
# Vistas para editar, eliminar y agregar direcciones
#===================================================
def editar_direccion(request, cliente_id, direccion_id):
    direccion = get_object_or_404(Direccion, id=direccion_id, cliente_id=cliente_id)
    if request.method == 'POST':
        form = DireccionForm(request.POST, instance=direccion)
        if form.is_valid():
            form.save()
            return redirect('editar_cliente', pk=cliente_id)
    else:
        form = DireccionForm(instance=direccion)

    return render(request, 'clientes/direccion_formulario.html', {
        'form': form,
        'cliente_id': cliente_id
    })

def eliminar_direccion(request, cliente_id, direccion_id):
    direccion = get_object_or_404(Direccion, id=direccion_id, cliente_id=cliente_id)
    if request.method == 'POST':
        direccion.delete()
        return redirect('editar_cliente', pk=cliente_id)
    return render(request, 'clientes/confirmar_eliminar_direccion.html', {
        'direccion': direccion,
        'cliente_id': cliente_id
    })

def agregar_direccion(request, cliente_id):
    cliente = get_object_or_404(Cliente, pk=cliente_id)
    titulo = "Nueva Dirección"
    if request.method == 'POST':
        form = DireccionForm(request.POST)
        if form.is_valid():
            nueva_direccion = form.save(commit=False)
            nueva_direccion.cliente = cliente
            nueva_direccion.save()
            return redirect('editar_cliente', pk=cliente_id)
    else:
        form = DireccionForm()

    return render(request, 'clientes/direccion_formulario.html', {
        'form': form,
        'cliente_id': cliente_id,
        'titulo': titulo,
    })

#===================================================
# Vistas para editar, eliminar y agregar direcciones
#===================================================
def guardar_direccion(request, cliente_id, direccion_id=None):
    cliente = get_object_or_404(Cliente, pk=cliente_id)
    if direccion_id:
        direccion = get_object_or_404(Direccion, pk=direccion_id, cliente=cliente)
        titulo = _("Editar Dirección")
    else:
        direccion = None
        titulo = "Nueva Dirección"
        print("prueba 1")

    if request.method == 'POST':
        form = DireccionForm(request.POST, instance=direccion)
        if form.is_valid():
            nueva_direccion = form.save(commit=False)
            nueva_direccion.cliente = cliente
            nueva_direccion.save()
            return redirect('editar_cliente', pk=cliente_id)
    else:
        form = DireccionForm(instance=direccion)
        print("prueba 2")

    #print(cliente_id)

    return render(request, 'clientes/direccion_formulario.html', {
        'form': form,
        'cliente_id': cliente_id,
        'titulo': titulo,
    })


#====================================
# Vistas para exportar a Excel/Pdf
#====================================
def exportar_clientes_excel(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Clientes'

    # Cabecera
    encabezados = ['ID', 'Nombre', 'Email', 'Teléfono', 'Direcciones']
    ws.append(encabezados)

    clientes = Cliente.objects.prefetch_related('direcciones').all()

    for cliente in clientes:
        # Unir direcciones en un solo string por cliente
        direcciones = cliente.direcciones.all()
        texto_direcciones = "\n".join([
            f"{d.calle}, {d.ciudad}, {d.region}, {d.codigo_postal}"
            for d in direcciones
        ])

        ws.append([
            cliente.id,
            cliente.nombre,
            cliente.email,
            cliente.telefono,
            texto_direcciones
        ])

    # Ajustar tamaño de columnas
    for col in ws.columns:
        max_length = 0
        columna = col[0].column
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        ajusta = get_column_letter(columna)
        ws.column_dimensions[ajusta].width = max_length + 2

    # Respuesta HTTP
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=clientes.xlsx'
    wb.save(response)
    return response


def exportar_clientes_pdf(request):
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename=clientes.pdf'

    p = canvas.Canvas(response, pagesize=letter)
    width, height = letter
    y = height - 40

    p.setFont("Helvetica-Bold", 14)
    p.drawString(50, y, "Lista de Clientes")
    y -= 30

    p.setFont("Helvetica", 10)

    clientes = Cliente.objects.prefetch_related('direcciones').all()

    for cliente in clientes:
        if y < 100:
            p.showPage()
            y = height - 40
            p.setFont("Helvetica-Bold", 14)
            p.drawString(50, y, "Lista de Clientes (continuación)")
            y -= 30
            p.setFont("Helvetica", 10)

        p.drawString(50, y, f"ID: {cliente.id}")
        p.drawString(100, y, f"Nombre: {cliente.nombre}")
        p.drawString(300, y, f"Correo: {cliente.email}")
        y -= 15
        p.drawString(100, y, f"Teléfono: {cliente.telefono}")
        y -= 15

        direcciones = cliente.direcciones.all()
        if direcciones:
            for direccion in direcciones:
                direccion_str = f"{direccion.calle}, {direccion.ciudad}, {direccion.region}, {direccion.codigo_postal}"
                p.drawString(120, y, f"Dirección: {direccion_str}")
                y -= 15
        else:
            p.drawString(120, y, "Dirección: (sin registrar)")
            y -= 15

        y -= 10  # Espacio entre clientes

    p.showPage()
    p.save()
    return response

